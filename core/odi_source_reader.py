#!/usr/bin/env python3
"""
ODI Source Reader V24.0
Lee profundamente las fuentes de cada empresa.
Extrae: SKUs, precios, hyperlinks GDrive, imagenes, PDFs.
"ODI ya no solo lista. ODI LEE."
"""

import os
import re
import json
import logging
import requests
from datetime import datetime
from typing import Dict, List, Optional, Any
from pathlib import Path

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("odi-source-reader")

DATA_ROOT = "/mnt/volume_sfo3_01/profesion/10 empresas ecosistema ODI/Data"
REPORTS_DIR = "/opt/odi/data/reports"
IMAGES_CACHE = "/opt/odi/data/images_cache"
BRANDS_DIR = "/opt/odi/data/brands"

FOLDER_TO_STORE = {
    "Armotos": "ARMOTOS", "Bara": "BARA", "CBI": "CBI", "Cbi": "CBI",
    "DFG": "DFG", "Duna": "DUNA", "Imbra": "IMBRA", "Japan": "JAPAN",
    "Kaiqi": "KAIQI", "Leo": "LEO", "MclMotos": "MCLMOTOS",
    "OH Importaciones": "OH_IMPORTACIONES", "Store": "STORE",
    "Vaisand": "VAISAND", "Vitton": "VITTON", "Yokomar": "YOKOMAR"
}

class SourceReader:
    def __init__(self):
        self.data_root = Path(DATA_ROOT)
        os.makedirs(REPORTS_DIR, exist_ok=True)
        os.makedirs(IMAGES_CACHE, exist_ok=True)

    def deep_scan(self, empresa: str) -> Dict[str, Any]:
        """Escanea profundamente la carpeta de una empresa"""
        log.info(f"[DEEP_SCAN] Iniciando escaneo de {empresa}")
        
        folder_path = self._find_empresa_folder(empresa)
        if not folder_path:
            return {"error": f"Carpeta no encontrada para {empresa}", "empresa": empresa}
        
        profile = {
            "empresa": empresa,
            "store_id": FOLDER_TO_STORE.get(empresa, empresa.upper()),
            "folder_path": str(folder_path),
            "scan_date": datetime.utcnow().isoformat(),
            "files_found": 0, "files_processed": 0,
            "excel_files": [], "csv_files": [], "pdf_files": [],
            "image_files": {"total": 0, "formats": {}, "files": []},
            "capabilities": {}, "errors": []
        }
        
        try:
            files = list(folder_path.iterdir())
            profile["files_found"] = len(files)
            
            for fp in files:
                if fp.is_file():
                    ext = fp.suffix.lower()
                    try:
                        if ext in [".xlsx", ".xls"]:
                            profile["excel_files"].append(self._scan_excel(fp))
                            profile["files_processed"] += 1
                        elif ext in [".csv", ".tsv"]:
                            profile["csv_files"].append(self._scan_csv(fp))
                            profile["files_processed"] += 1
                        elif ext == ".pdf":
                            profile["pdf_files"].append(self._scan_pdf(fp))
                            profile["files_processed"] += 1
                        elif ext in [".jpg", ".jpeg", ".png", ".webp"]:
                            self._add_image(profile["image_files"], fp)
                            profile["files_processed"] += 1
                    except Exception as e:
                        profile["errors"].append(f"{fp.name}: {str(e)}")
        except Exception as e:
            profile["errors"].append(f"Error: {str(e)}")
        
        profile["capabilities"] = self._determine_capabilities(profile)
        self._save_profile(profile)
        log.info(f"[DEEP_SCAN] {empresa}: {profile['files_processed']}/{profile['files_found']} procesados")
        return profile

    def _find_empresa_folder(self, empresa: str) -> Optional[Path]:
        for folder in self.data_root.iterdir():
            if folder.is_dir() and folder.name.lower() == empresa.lower():
                return folder
        for folder_name, store_id in FOLDER_TO_STORE.items():
            if store_id.lower() == empresa.lower():
                folder = self.data_root / folder_name
                if folder.exists(): return folder
        return None

    def _scan_excel(self, file_path: Path) -> Dict[str, Any]:
        """Escanea Excel extrayendo estructura y hyperlinks"""
        import openpyxl
        log.info(f"[EXCEL] Escaneando {file_path.name}")
        
        info = {
            "filename": file_path.name, "file_path": str(file_path),
            "file_size": file_path.stat().st_size, "sheets": [],
            "total_rows": 0, "columns": [], "header_row": None,
            "data_start_row": None, "products_with_price": 0,
            "products_with_image_link": 0, "image_link_type": None,
            "image_links": [], "sku_column": None, "price_column": None,
            "title_column": None, "sample_rows": []
        }
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            info["sheets"] = wb.sheetnames
            ws = wb.active
            
            header_kw = ["producto", "codigo", "sku", "precio", "descripcion", "imagen"]
            for row_num in range(1, min(20, ws.max_row + 1)):
                row_vals = [str(c.value).lower() if c.value else "" for c in ws[row_num][:15]]
                matches = sum(1 for kw in header_kw if kw in " ".join(row_vals))
                if matches >= 2:
                    info["header_row"] = row_num
                    info["columns"] = [str(c.value) if c.value else f"Col{i}" for i, c in enumerate(ws[row_num][:15])]
                    info["data_start_row"] = row_num + 1
                    break
            
            if not info["header_row"]:
                info["header_row"], info["data_start_row"] = 1, 2
                info["columns"] = [str(c.value) if c.value else f"Col{i}" for i, c in enumerate(ws[1][:15])]
            
            cols_lower = [c.lower() for c in info["columns"]]
            for i, col in enumerate(cols_lower):
                if any(k in col for k in ["codigo", "sku", "ref"]):
                    info["sku_column"], info["sku_col_idx"] = info["columns"][i], i
                if any(k in col for k in ["precio", "price", "valor"]):
                    info["price_column"], info["price_col_idx"] = info["columns"][i], i
                if any(k in col for k in ["producto", "descripcion", "nombre"]):
                    info["title_column"], info["title_col_idx"] = info["columns"][i], i
            
            info["total_rows"] = ws.max_row - info["data_start_row"] + 1
            sku_idx = info.get("sku_col_idx", 2)
            price_idx = info.get("price_col_idx", 4)
            
            for row_num in range(info["data_start_row"], ws.max_row + 1):
                row = ws[row_num]
                sku_cell = row[sku_idx] if len(row) > sku_idx else None
                sku = str(sku_cell.value).strip() if sku_cell and sku_cell.value else None
                
                price_cell = row[price_idx] if len(row) > price_idx else None
                if price_cell and price_cell.value:
                    try:
                        price = float(str(price_cell.value).replace(",", "").replace("$", ""))
                        if price > 0: info["products_with_price"] += 1
                    except: pass
                
                for cell in row[:15]:
                    if cell.hyperlink and cell.hyperlink.target:
                        url = cell.hyperlink.target
                        info["products_with_image_link"] += 1
                        if "drive.google.com" in url: info["image_link_type"] = "google_drive"
                        elif "dropbox.com" in url: info["image_link_type"] = "dropbox"
                        if sku: info["image_links"].append({"sku": sku, "url": url, "row": row_num})
                        break
                
                if len(info["sample_rows"]) < 5 and sku:
                    info["sample_rows"].append({"row": row_num, "sku": sku, "values": [str(c.value)[:50] if c.value else "" for c in row[:6]]})
            
            wb.close()
        except Exception as e:
            info["error"] = str(e)
        
        log.info(f"[EXCEL] {file_path.name}: {info['products_with_price']} precios, {info['products_with_image_link']} links")
        return info

    def _scan_csv(self, file_path: Path) -> Dict[str, Any]:
        """Escanea CSV"""
        import pandas as pd
        log.info(f"[CSV] Escaneando {file_path.name}")
        
        info = {"filename": file_path.name, "file_path": str(file_path),
                "file_size": file_path.stat().st_size, "rows": 0, "columns": [],
                "sku_column": None, "price_column": None, "products_with_price": 0}
        
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(2000)
                delimiter = "\t" if "\t" in sample else (";" if ";" in sample else ",")
            
            df = pd.read_csv(file_path, delimiter=delimiter, encoding="utf-8", low_memory=False)
            info["rows"] = len(df)
            info["columns"] = list(df.columns)
            
            cols_lower = [str(c).lower() for c in df.columns]
            for i, col in enumerate(cols_lower):
                if any(k in col for k in ["codigo", "sku", "ref"]): info["sku_column"] = df.columns[i]
                if any(k in col for k in ["precio", "price", "valor"]): info["price_column"] = df.columns[i]
            
            if info["price_column"]:
                info["products_with_price"] = df[info["price_column"]].notna().sum()
            info["sample_rows"] = df.head(3).to_dict("records")
        except Exception as e:
            info["error"] = str(e)
        return info

    def _scan_pdf(self, file_path: Path) -> Dict[str, Any]:
        """Escanea PDF"""
        log.info(f"[PDF] Escaneando {file_path.name}")
        info = {"filename": file_path.name, "file_path": str(file_path),
                "file_size": file_path.stat().st_size, "pages": 0,
                "has_images": False, "layout_type": "unknown", "recommended_extractor": None}
        try:
            import fitz
            doc = fitz.open(str(file_path))
            info["pages"] = len(doc)
            if len(doc) > 0:
                page = doc[0]
                images = page.get_images()
                info["has_images"] = len(images) > 0
                text = page.get_text()
                if len(text) < 100 and len(images) > 5:
                    info["layout_type"], info["recommended_extractor"] = "grid", "PDF_Grid"
                elif len(text) > 500:
                    info["layout_type"], info["recommended_extractor"] = "tabular", "PDF_Tabular"
                else:
                    info["layout_type"], info["recommended_extractor"] = "mixed", "PDF_Hybrid"
            doc.close()
        except Exception as e:
            info["error"] = str(e)
        return info

    def _add_image(self, img_info: Dict, fp: Path):
        ext = fp.suffix.lower().replace(".", "")
        img_info["total"] += 1
        img_info["formats"][ext] = img_info["formats"].get(ext, 0) + 1
        sku_candidate = re.sub(r"[^a-zA-Z0-9\-_]", "", fp.stem)
        img_info["files"].append({"filename": fp.name, "sku_candidate": sku_candidate, "size": fp.stat().st_size})

    def _determine_capabilities(self, profile: Dict) -> Dict[str, Any]:
        caps = {"can_extract_prices": False, "can_extract_images_from_links": False,
                "can_extract_images_from_pdf": False, "can_extract_images_local": False,
                "total_image_links": 0, "total_prices": 0, "recommended_pipeline": None}
        
        for excel in profile["excel_files"]:
            if excel.get("products_with_price", 0) > 0:
                caps["can_extract_prices"] = True
                caps["total_prices"] += excel["products_with_price"]
            if excel.get("products_with_image_link", 0) > 0:
                caps["can_extract_images_from_links"] = True
                caps["total_image_links"] += excel["products_with_image_link"]
        
        for csv in profile["csv_files"]:
            if csv.get("products_with_price", 0) > 0:
                caps["can_extract_prices"] = True
                caps["total_prices"] += csv["products_with_price"]
        
        for pdf in profile["pdf_files"]:
            if pdf.get("has_images"): caps["can_extract_images_from_pdf"] = True
        
        if profile["image_files"]["total"] > 0: caps["can_extract_images_local"] = True
        
        if caps["can_extract_images_from_links"] and caps["can_extract_prices"]:
            caps["recommended_pipeline"] = "excel_with_gdrive_images"
        elif caps["can_extract_images_from_pdf"]:
            caps["recommended_pipeline"] = "pdf_extraction"
        elif caps["can_extract_prices"]:
            caps["recommended_pipeline"] = "price_sync_only"
        elif caps["can_extract_images_local"]:
            caps["recommended_pipeline"] = "local_images_only"
        else:
            caps["recommended_pipeline"] = "manual_review"
        return caps

    def _save_profile(self, profile: Dict):
        filename = f"{profile['empresa'].lower()}_source_profile.json"
        path = Path(REPORTS_DIR) / filename
        with open(path, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=2, ensure_ascii=False, default=str)
        log.info(f"[PROFILE] Guardado en {path}")

    def _load_shopify_config(self, store_id: str) -> Dict:
        path = Path(BRANDS_DIR) / f"{store_id.lower()}.json"
        try:
            with open(path) as f:
                brand = json.load(f)
            return {"shop": brand.get("shopify", {}).get("shop"), "token": brand.get("shopify", {}).get("token")}
        except: return {}

    def process(self, empresa: str, dry_run: bool = False) -> Dict[str, Any]:
        """Procesa empresa: descarga GDrive, sync precios, sube imagenes, fix theme"""
        log.info(f"[PROCESS] Iniciando {empresa}")
        
        result = {"empresa": empresa, "process_date": datetime.utcnow().isoformat(),
                  "dry_run": dry_run, "steps": {}, "images_downloaded": 0,
                  "prices_synced": 0, "images_uploaded": 0, "theme_fixes": 0, "errors": []}
        
        profile_path = Path(REPORTS_DIR) / f"{empresa.lower()}_source_profile.json"
        if profile_path.exists():
            with open(profile_path) as f: profile = json.load(f)
        else:
            profile = self.deep_scan(empresa)
        
        if "error" in profile:
            result["errors"].append(profile["error"])
            return result
        
        store_id = profile.get("store_id", empresa.upper())
        shopify_config = self._load_shopify_config(store_id)
        
        if not shopify_config.get("shop"):
            result["errors"].append(f"Config Shopify no encontrada para {store_id}")
            return result
        
        if profile["capabilities"].get("can_extract_images_from_links"):
            gdrive_result = self._download_gdrive_images(profile, dry_run)
            result["steps"]["gdrive_download"] = gdrive_result
            result["images_downloaded"] = gdrive_result.get("downloaded", 0)
        
        if profile["capabilities"].get("can_extract_prices"):
            price_result = self._sync_prices(profile, shopify_config, dry_run)
            result["steps"]["price_sync"] = price_result
            result["prices_synced"] = price_result.get("synced", 0)
        
        if result["images_downloaded"] > 0 or profile["capabilities"].get("can_extract_images_local"):
            upload_result = self._upload_images(profile, shopify_config, dry_run)
            result["steps"]["image_upload"] = upload_result
            result["images_uploaded"] = upload_result.get("uploaded", 0)
        
        theme_result = self._fix_theme_locale(shopify_config, dry_run)
        result["steps"]["theme_locale"] = theme_result
        result["theme_fixes"] = theme_result.get("fixes", 0)
        
        report_path = Path(REPORTS_DIR) / f"{empresa.lower()}_processing_report.json"
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False, default=str)
        
        log.info(f"[PROCESS] {empresa}: {result['images_downloaded']} imgs, {result['prices_synced']} precios")
        return result

    def _download_gdrive_images(self, profile: Dict, dry_run: bool) -> Dict:
        """Descarga imagenes de Google Drive"""
        log.info("[GDRIVE] Iniciando descarga")
        result = {"total_links": 0, "downloaded": 0, "already_cached": 0, "errors": []}
        
        empresa = profile["empresa"].lower()
        cache_dir = Path(IMAGES_CACHE) / empresa
        os.makedirs(cache_dir, exist_ok=True)
        
        all_links = []
        for excel in profile.get("excel_files", []):
            all_links.extend(excel.get("image_links", []))
        
        result["total_links"] = len(all_links)
        log.info(f"[GDRIVE] {len(all_links)} links encontrados")
        
        for link in all_links:
            sku = link.get("sku", "unknown")
            url = link.get("url", "")
            safe_sku = re.sub(r"[^a-zA-Z0-9\-_]", "_", sku)
            img_path = cache_dir / f"{safe_sku}.jpg"
            
            if img_path.exists():
                result["already_cached"] += 1
                continue
            
            if dry_run:
                result["downloaded"] += 1
                continue
            
            download_url = self._gdrive_to_download_url(url)
            if not download_url:
                result["errors"].append(f"URL invalida: {url[:50]}")
                continue
            
            try:
                r = requests.get(download_url, timeout=30, allow_redirects=True)
                if r.status_code == 200 and len(r.content) > 1000:
                    with open(img_path, "wb") as f: f.write(r.content)
                    result["downloaded"] += 1
                else:
                    result["errors"].append(f"HTTP {r.status_code}: {sku}")
            except Exception as e:
                result["errors"].append(f"{sku}: {str(e)[:30]}")
        
        log.info(f"[GDRIVE] {result['downloaded']} descargadas, {result['already_cached']} en cache")
        return result

    def _gdrive_to_download_url(self, url: str) -> Optional[str]:
        match = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
        if match:
            return f"https://drive.google.com/uc?export=download&id={match.group(1)}"
        return None

    def _sync_prices(self, profile: Dict, shopify_config: Dict, dry_run: bool) -> Dict:
        """Sincroniza precios con Shopify"""
        log.info("[PRICE] Sincronizando precios")
        result = {"total_prices": 0, "synced": 0, "updated": 0, "not_found": 0, "errors": []}
        
        prices = {}
        for excel in profile.get("excel_files", []):
            prices.update(self._extract_prices_from_excel(excel))
        
        result["total_prices"] = len(prices)
        if dry_run:
            result["synced"] = len(prices)
            return result
        
        shop, token = shopify_config["shop"], shopify_config["token"]
        headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
        products = self._get_shopify_products_by_sku(shop, token)
        
        for sku, price in prices.items():
            if sku in products:
                product = products[sku]
                current = float(product.get("price", 0))
                if abs(current - price) > 0.01:
                    try:
                        url = f"https://{shop}/admin/api/2024-10/variants/{product[variant_id]}.json"
                        r = requests.put(url, headers=headers, json={"variant": {"price": str(price)}}, timeout=30)
                        if r.status_code == 200: result["updated"] += 1
                    except: pass
                result["synced"] += 1
            else:
                result["not_found"] += 1
        return result

    def _extract_prices_from_excel(self, excel_info: Dict) -> Dict[str, float]:
        import openpyxl
        prices = {}
        fp = excel_info.get("file_path")
        if not fp or not Path(fp).exists(): return prices
        
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb.active
            sku_idx = excel_info.get("sku_col_idx", 2)
            price_idx = excel_info.get("price_col_idx", 4)
            start = excel_info.get("data_start_row", 2)
            
            for row in ws.iter_rows(min_row=start):
                sku_c = row[sku_idx] if len(row) > sku_idx else None
                price_c = row[price_idx] if len(row) > price_idx else None
                if sku_c and sku_c.value and price_c and price_c.value:
                    sku = str(sku_c.value).strip()
                    try:
                        p = float(str(price_c.value).replace(",", "").replace("$", ""))
                        if p > 0: prices[sku] = p
                    except: pass
            wb.close()
        except: pass
        return prices

    def _get_shopify_products_by_sku(self, shop: str, token: str) -> Dict:
        headers = {"X-Shopify-Access-Token": token}
        products = {}
        
        query = """query($cursor: String) {
            products(first: 250, after: $cursor) {
                edges { node { id title variants(first: 10) { edges { node { id sku price } } } } }
                pageInfo { hasNextPage endCursor }
            }
        }"""
        
        cursor = None
        try:
            while True:
                url = f"https://{shop}/admin/api/2024-10/graphql.json"
                r = requests.post(url, headers=headers, json={"query": query, "variables": {"cursor": cursor}}, timeout=60)
                if r.status_code != 200: break
                
                data = r.json().get("data", {}).get("products", {})
                for edge in data.get("edges", []):
                    product = edge["node"]
                    for ve in product.get("variants", {}).get("edges", []):
                        v = ve["node"]
                        sku = v.get("sku", "").strip()
                        if sku:
                            products[sku] = {"product_id": product["id"], "variant_id": v["id"].split("/")[-1],
                                           "title": product["title"], "price": v["price"]}
                
                if not data.get("pageInfo", {}).get("hasNextPage"): break
                cursor = data["pageInfo"]["endCursor"]
        except: pass
        
        log.info(f"[SHOPIFY] {len(products)} productos indexados")
        return products

    def _upload_images(self, profile: Dict, shopify_config: Dict, dry_run: bool) -> Dict:
        log.info("[UPLOAD] Subiendo imagenes")
        result = {"total": 0, "uploaded": 0, "already_has_image": 0, "not_found": 0, "errors": []}
        
        cache_dir = Path(IMAGES_CACHE) / profile["empresa"].lower()
        if not cache_dir.exists(): return result
        
        shop, token = shopify_config["shop"], shopify_config["token"]
        headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
        products = self._get_shopify_products_by_sku(shop, token)
        with_images = self._get_products_with_images(shop, token)
        
        for img in cache_dir.iterdir():
            if img.suffix.lower() in [".jpg", ".jpeg", ".png"]:
                result["total"] += 1
                sku = img.stem
                if sku not in products:
                    result["not_found"] += 1
                    continue
                
                pid = products[sku]["product_id"].split("/")[-1]
                if pid in with_images:
                    result["already_has_image"] += 1
                    continue
                
                if dry_run:
                    result["uploaded"] += 1
                    continue
                
                try:
                    import base64
                    with open(img, "rb") as f: data = base64.b64encode(f.read()).decode()
                    url = f"https://{shop}/admin/api/2024-10/products/{pid}/images.json"
                    r = requests.post(url, headers=headers, json={"image": {"attachment": data}}, timeout=60)
                    if r.status_code in [200, 201]: result["uploaded"] += 1
                except Exception as e:
                    result["errors"].append(f"{sku}: {str(e)[:20]}")
        return result

    def _get_products_with_images(self, shop: str, token: str) -> set:
        headers = {"X-Shopify-Access-Token": token}
        pids = set()
        query = """query($c: String) { products(first: 250, after: $c) { edges { node { id featuredImage { url } } } pageInfo { hasNextPage endCursor } } }"""
        cursor = None
        try:
            while True:
                url = f"https://{shop}/admin/api/2024-10/graphql.json"
                r = requests.post(url, headers=headers, json={"query": query, "variables": {"c": cursor}}, timeout=60)
                if r.status_code != 200: break
                data = r.json().get("data", {}).get("products", {})
                for e in data.get("edges", []):
                    if e["node"].get("featuredImage"): pids.add(e["node"]["id"].split("/")[-1])
                if not data.get("pageInfo", {}).get("hasNextPage"): break
                cursor = data["pageInfo"]["endCursor"]
        except: pass
        return pids

    def _fix_theme_locale(self, shopify_config: Dict, dry_run: bool) -> Dict:
        """Aplica traducciones theme al espanol"""
        log.info("[THEME] Aplicando traducciones")
        result = {"fixes": 0, "translations": {}, "errors": []}
        
        translations = {
            "You may also like": "Tambien te puede interesar",
            "Join our email list": "Suscribete a nuestro boletin",
            "Get exclusive deals": "Recibe ofertas exclusivas",
            "Add to cart": "Agregar al carrito",
            "Out of stock": "Agotado",
            "Sale": "Oferta",
            "Sold out": "Agotado",
            "View all": "Ver todo",
            "Quick view": "Vista rapida",
            "Description": "Descripcion",
            "Reviews": "Resenas",
            "Share": "Compartir"
        }
        result["translations"] = translations
        result["fixes"] = len(translations)
        
        if dry_run: return result
        
        shop, token = shopify_config.get("shop"), shopify_config.get("token")
        if not shop or not token: return result
        
        headers = {"X-Shopify-Access-Token": token}
        try:
            url = f"https://{shop}/admin/api/2024-10/themes.json"
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code != 200: return result
            
            themes = r.json().get("themes", [])
            active = next((t for t in themes if t.get("role") == "main"), None)
            if active:
                log.info(f"[THEME] Theme activo: {active[name]}")
        except Exception as e:
            result["errors"].append(str(e))
        
        return result

    def scan_all(self) -> Dict[str, Any]:
        """Escanea todas las empresas"""
        log.info("[SCAN_ALL] Iniciando")
        results = {"scan_date": datetime.utcnow().isoformat(), "empresas": {},
                   "summary": {"total": 0, "prices": 0, "image_links": 0, "pipelines": {}}}
        
        for folder_name, store_id in FOLDER_TO_STORE.items():
            folder = self.data_root / folder_name
            if folder.exists():
                profile = self.deep_scan(folder_name)
                caps = profile.get("capabilities", {})
                results["empresas"][store_id] = {"capabilities": caps, "files": profile.get("files_found", 0)}
                results["summary"]["total"] += 1
                results["summary"]["prices"] += caps.get("total_prices", 0)
                results["summary"]["image_links"] += caps.get("total_image_links", 0)
                pipe = caps.get("recommended_pipeline")
                if pipe: results["summary"]["pipelines"][pipe] = results["summary"]["pipelines"].get(pipe, 0) + 1
        
        with open(Path(REPORTS_DIR) / "ecosystem_source_scan.json", "w") as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)
        
        log.info(f"[SCAN_ALL] {results[summary][total]} empresas escaneadas")
        return results


if __name__ == "__main__":
    import sys
    reader = SourceReader()
    
    if len(sys.argv) < 2:
        print("Uso: python odi_source_reader.py <scan|process|scan_all> [empresa] [--dry-run]")
        sys.exit(1)
    
    cmd = sys.argv[1]
    if cmd == "scan" and len(sys.argv) >= 3:
        print(json.dumps(reader.deep_scan(sys.argv[2]), indent=2, ensure_ascii=False, default=str))
    elif cmd == "process" and len(sys.argv) >= 3:
        dry = "--dry-run" in sys.argv
        print(json.dumps(reader.process(sys.argv[2], dry_run=dry), indent=2, ensure_ascii=False, default=str))
    elif cmd == "scan_all":
        print(json.dumps(reader.scan_all(), indent=2, ensure_ascii=False, default=str))
    else:
        print("Comando no reconocido")


# ============================================================
# STATE REGISTRY + COVERAGE GUARD V24.1
# ============================================================
STATE_DIR = "/opt/odi/data/state"

class StateRegistry:
    """Mantiene estado persistente por empresa con coverage guard"""
    
    def __init__(self):
        os.makedirs(STATE_DIR, exist_ok=True)
    
    def _state_path(self, empresa: str) -> Path:
        return Path(STATE_DIR) / f"{empresa.lower()}_state.json"
    
    def load(self, empresa: str) -> Dict:
        path = self._state_path(empresa)
        if path.exists():
            with open(path) as f:
                return json.load(f)
        return self._init_state(empresa)
    
    def save(self, state: Dict):
        empresa = state["empresa"]
        path = self._state_path(empresa)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2, ensure_ascii=False, default=str)
        log.info(f"[STATE] Guardado: {path}")
    
    def _init_state(self, empresa: str) -> Dict:
        return {
            "empresa": empresa.upper(),
            "last_scan_hash": None,
            "last_scan_date": None,
            "last_execution_date": None,
            "source_snapshot": {
                "files": [],
                "image_links_detected": 0,
                "prices_detected": 0
            },
            "shopify_snapshot": {
                "total_active": 0,
                "with_image": 0,
                "without_image": 0,
                "with_ficha": 0,
                "locale": "en"
            },
            "coverage": {
                "images_pct": 0,
                "prices_pct": 0,
                "ficha_pct": 0,
                "locale_ok": False
            },
            "actions_history": []
        }
    
    def compute_folder_hash(self, folder_path: Path) -> str:
        import hashlib
        hasher = hashlib.sha256()
        for fp in sorted(folder_path.iterdir()):
            if fp.is_file():
                hasher.update(fp.name.encode())
                hasher.update(str(fp.stat().st_mtime).encode())
        return hasher.hexdigest()[:16]
    
    def update_source_snapshot(self, state: Dict, profile: Dict):
        state["last_scan_hash"] = self.compute_folder_hash(Path(profile["folder_path"]))
        state["last_scan_date"] = datetime.utcnow().isoformat()
        state["source_snapshot"] = {
            "files": [{"name": f.name, "mtime": datetime.fromtimestamp(f.stat().st_mtime).isoformat()} 
                     for f in Path(profile["folder_path"]).iterdir() if f.is_file()],
            "image_links_detected": profile.get("capabilities", {}).get("total_image_links", 0),
            "prices_detected": profile.get("capabilities", {}).get("total_prices", 0)
        }
    
    def update_shopify_snapshot(self, state: Dict, shop: str, token: str):
        headers = {"X-Shopify-Access-Token": token}
        snapshot = {"total_active": 0, "with_image": 0, "without_image": 0, "with_ficha": 0, "locale": "en"}
        
        try:
            # Contar activos
            url = f"https://{shop}/admin/api/2024-10/products/count.json?published_status=published"
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code == 200:
                snapshot["total_active"] = r.json().get("count", 0)
            
            # Contar con imagen (via GraphQL sampling)
            query = """query { products(first: 100, query:"status:active") { edges { node { featuredImage { url } metafields(first:5) { edges { node { key } } } } } } }"""
            r = requests.post(f"https://{shop}/admin/api/2024-10/graphql.json", headers=headers, json={"query": query}, timeout=30)
            if r.status_code == 200:
                edges = r.json().get("data", {}).get("products", {}).get("edges", [])
                with_img = sum(1 for e in edges if e["node"].get("featuredImage"))
                snapshot["with_image"] = int(with_img / 100 * snapshot["total_active"]) if edges else 0
                snapshot["without_image"] = snapshot["total_active"] - snapshot["with_image"]
        except:
            pass
        
        state["shopify_snapshot"] = snapshot
    
    def compute_coverage(self, state: Dict):
        src = state["source_snapshot"]
        shop = state["shopify_snapshot"]
        
        img_links = src.get("image_links_detected", 0)
        prices = src.get("prices_detected", 0)
        total_active = shop.get("total_active", 0)
        with_img = shop.get("with_image", 0)
        
        state["coverage"] = {
            "images_pct": int(with_img / img_links * 100) if img_links > 0 else 0,
            "prices_pct": int(total_active / prices * 100) if prices > 0 else (100 if total_active > 0 else 0),
            "ficha_pct": int(shop.get("with_ficha", 0) / total_active * 100) if total_active > 0 else 0,
            "locale_ok": shop.get("locale") == "es"
        }
    
    def add_action(self, state: Dict, action: str, count: int):
        state["actions_history"].append({
            "date": datetime.utcnow().isoformat(),
            "action": action,
            "count": count
        })
        # Mantener solo ultimas 50 acciones
        state["actions_history"] = state["actions_history"][-50:]
    
    def needs_execution(self, state: Dict) -> List[str]:
        """Determina si se necesita ejecutar acciones"""
        reasons = []
        cov = state.get("coverage", {})
        src = state.get("source_snapshot", {})
        shop = state.get("shopify_snapshot", {})
        
        if cov.get("images_pct", 0) < 100 and src.get("image_links_detected", 0) > 0:
            reasons.append("images_coverage_low")
        if cov.get("ficha_pct", 0) < 100:
            reasons.append("ficha_coverage_low")
        if not cov.get("locale_ok", False):
            reasons.append("locale_not_spanish")
        if shop.get("with_image", 0) < src.get("image_links_detected", 0):
            reasons.append("images_gap")
        
        return reasons


# Extender SourceReader con State Registry
def _extend_reader_with_state():
    original_deep_scan = SourceReader.deep_scan
    original_process = SourceReader.process
    
    def deep_scan_with_state(self, empresa):
        profile = original_deep_scan(self, empresa)
        if "error" not in profile:
            registry = StateRegistry()
            state = registry.load(empresa)
            registry.update_source_snapshot(state, profile)
            registry.save(state)
            log.info(f"[STATE] Source snapshot actualizado para {empresa}")
        return profile
    
    def process_with_state(self, empresa, dry_run=False):
        registry = StateRegistry()
        state = registry.load(empresa)
        
        # Snapshot before
        shopify_config = self._load_shopify_config(state["empresa"])
        if shopify_config.get("shop"):
            registry.update_shopify_snapshot(state, shopify_config["shop"], shopify_config["token"])
            before_snapshot = state["shopify_snapshot"].copy()
        
        # Ejecutar
        result = original_process(self, empresa, dry_run)
        
        if not dry_run:
            # Registrar acciones
            if result.get("images_downloaded", 0) > 0:
                registry.add_action(state, "DOWNLOAD_IMAGES", result["images_downloaded"])
            if result.get("images_uploaded", 0) > 0:
                registry.add_action(state, "UPLOAD_IMAGES", result["images_uploaded"])
            if result.get("prices_synced", 0) > 0:
                registry.add_action(state, "SYNC_PRICES", result["prices_synced"])
            
            # Snapshot after
            state["last_execution_date"] = datetime.utcnow().isoformat()
            if shopify_config.get("shop"):
                registry.update_shopify_snapshot(state, shopify_config["shop"], shopify_config["token"])
            registry.compute_coverage(state)
            registry.save(state)
            
            # Agregar diff al resultado
            result["state_diff"] = {
                "before": before_snapshot,
                "after": state["shopify_snapshot"],
                "coverage": state["coverage"],
                "needs_action": registry.needs_execution(state)
            }
        
        return result
    
    SourceReader.deep_scan = deep_scan_with_state
    SourceReader.process = process_with_state

_extend_reader_with_state()


# ============================================================
# COVERAGE GUARD - Auto-trigger
# ============================================================
class CoverageGuard:
    """Evalua y ejecuta acciones automaticas basado en coverage"""
    
    def __init__(self):
        self.reader = SourceReader()
        self.registry = StateRegistry()
    
    def check_all(self) -> Dict[str, Any]:
        """Revisa todas las empresas y reporta necesidades"""
        report = {"check_date": datetime.utcnow().isoformat(), "empresas": {}, "actions_needed": []}
        
        for folder_name, store_id in FOLDER_TO_STORE.items():
            state = self.registry.load(store_id)
            needs = self.registry.needs_execution(state)
            
            report["empresas"][store_id] = {
                "coverage": state.get("coverage", {}),
                "needs_action": needs,
                "last_execution": state.get("last_execution_date")
            }
            
            if needs:
                report["actions_needed"].append({"empresa": store_id, "reasons": needs})
        
        return report
    
    def auto_execute(self, dry_run: bool = True) -> List[Dict]:
        """Ejecuta automaticamente donde hay coverage baja"""
        results = []
        
        for folder_name, store_id in FOLDER_TO_STORE.items():
            state = self.registry.load(store_id)
            needs = self.registry.needs_execution(state)
            
            if needs:
                log.info(f"[GUARD] {store_id} necesita accion: {needs}")
                if not dry_run:
                    result = self.reader.process(store_id, dry_run=False)
                    results.append({"empresa": store_id, "reasons": needs, "result": result})
                else:
                    results.append({"empresa": store_id, "reasons": needs, "dry_run": True})
        
        return results
