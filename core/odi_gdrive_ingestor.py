#!/usr/bin/env python3
"""
ODI V24 - Google Drive Ingestor
Módulo permanente para ingestión de imágenes desde Excel con links GDrive.
Reutilizable para cualquier empresa del ecosistema ODI.
"""
import json
import os
import re
import requests
import time
import base64
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

BRANDS_DIR = "/opt/odi/data/brands"
CACHE_DIR = "/opt/odi/data/image_cache"
SOURCE_DIR = "/opt/odi/data/source"
REPORTS_DIR = "/opt/odi/data/reports"

# Validaciones
MIN_IMAGE_SIZE_KB = 5
VALID_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
RATE_LIMIT_SECONDS = 0.5


class GDriveIngestor:
    """
    Ingesta imágenes desde Google Drive y sincroniza con Shopify.
    Patrón ODI reutilizable para cualquier empresa.
    """

    def __init__(self, empresa: str):
        self.empresa = empresa.upper()
        self.cache_dir = os.path.join(CACHE_DIR, f"{self.empresa}_GDRIVE")
        self.config = self._load_config()
        os.makedirs(self.cache_dir, exist_ok=True)
        os.makedirs(REPORTS_DIR, exist_ok=True)

    def _load_config(self) -> Optional[Dict]:
        """Cargar configuración Shopify de la empresa."""
        path = os.path.join(BRANDS_DIR, f"{self.empresa.lower()}.json")
        if not os.path.exists(path):
            return None
        with open(path) as f:
            data = json.load(f)
            return data.get("shopify", data)

    @staticmethod
    def extract_gdrive_id(url: str) -> Optional[str]:
        """
        Extraer FILE_ID de URL de Google Drive.
        Soporta formatos:
        - https://drive.google.com/file/d/{ID}/view?usp=sharing
        - https://drive.google.com/open?id={ID}
        - https://drive.google.com/uc?id={ID}
        """
        if not url or not isinstance(url, str):
            return None

        patterns = [
            r'/file/d/([a-zA-Z0-9_-]+)',
            r'[?&]id=([a-zA-Z0-9_-]+)',
            r'/d/([a-zA-Z0-9_-]+)',
        ]

        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                return match.group(1)
        return None

    @staticmethod
    def gdrive_url_to_download(url: str) -> Optional[str]:
        """Convertir URL de GDrive a URL de descarga directa."""
        file_id = GDriveIngestor.extract_gdrive_id(url)
        if not file_id:
            return None
        return f"https://drive.google.com/uc?export=download&id={file_id}"

    def extract_links_from_excel(self, filepath: str) -> List[Dict]:
        """
        Extraer links de GDrive desde Excel.
        Busca columnas: SKU/Código, URL/Imagen/Link, Título/Nombre, Precio
        Retorna: [{sku, gdrive_url, title, price}, ...]
        """
        if not os.path.exists(filepath):
            print(f"[ERROR] Archivo no encontrado: {filepath}")
            return []

        results = []
        ext = os.path.splitext(filepath)[1].lower()

        if ext == '.xlsx' and openpyxl:
            results = self._extract_xlsx_openpyxl(filepath)
        elif pd is not None:
            results = self._extract_pandas(filepath)
        else:
            print("[ERROR] No hay librería disponible para leer Excel")
            return []

        print(f"[EXTRACT] {len(results)} registros con GDrive URL extraídos de {filepath}")
        return results

    def _extract_xlsx_openpyxl(self, filepath: str) -> List[Dict]:
        """Extraer usando openpyxl."""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        results = []

        # Detectar columnas por nombre
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            val = str(cell.value or '').lower().strip()
            if any(k in val for k in ['sku', 'codigo', 'código', 'ref']):
                headers['sku'] = col_idx
            elif any(k in val for k in ['url', 'imagen', 'link', 'drive', 'foto']):
                headers['url'] = col_idx
            elif any(k in val for k in ['titulo', 'título', 'nombre', 'descripcion', 'descripción']):
                headers['title'] = col_idx
            elif any(k in val for k in ['precio', 'price', 'valor']):
                headers['price'] = col_idx

        if 'sku' not in headers or 'url' not in headers:
            print(f"[WARN] Columnas requeridas no encontradas. Headers: {headers}")
            # Intentar por posición: A=SKU, B=Titulo, C=Precio, D=URL
            headers = {'sku': 1, 'title': 2, 'price': 3, 'url': 4}

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                sku = str(row[headers['sku'] - 1] or '').strip()
                url = str(row[headers.get('url', 4) - 1] or '').strip()
                title = str(row[headers.get('title', 2) - 1] or '').strip() if headers.get('title') else ''
                price = row[headers.get('price', 3) - 1] if headers.get('price') else None

                if sku and url and 'drive.google.com' in url:
                    results.append({
                        'sku': sku,
                        'gdrive_url': url,
                        'title': title,
                        'price': float(price) if price else None
                    })
            except (IndexError, ValueError, TypeError):
                continue

        return results

    def _extract_pandas(self, filepath: str) -> List[Dict]:
        """Extraer usando pandas."""
        df = pd.read_excel(filepath)
        results = []

        # Normalizar nombres de columnas
        df.columns = [str(c).lower().strip() for c in df.columns]

        # Detectar columnas
        sku_col = next((c for c in df.columns if any(k in c for k in ['sku', 'codigo', 'código', 'ref'])), None)
        url_col = next((c for c in df.columns if any(k in c for k in ['url', 'imagen', 'link', 'drive', 'foto'])), None)
        title_col = next((c for c in df.columns if any(k in c for k in ['titulo', 'título', 'nombre', 'descripcion'])), None)
        price_col = next((c for c in df.columns if any(k in c for k in ['precio', 'price', 'valor'])), None)

        if not sku_col or not url_col:
            print(f"[WARN] Columnas no detectadas. Usando posiciones por defecto.")
            cols = df.columns.tolist()
            sku_col = cols[0] if len(cols) > 0 else None
            url_col = cols[3] if len(cols) > 3 else None
            title_col = cols[1] if len(cols) > 1 else None
            price_col = cols[2] if len(cols) > 2 else None

        for _, row in df.iterrows():
            try:
                sku = str(row.get(sku_col, '') or '').strip()
                url = str(row.get(url_col, '') or '').strip()
                title = str(row.get(title_col, '') or '').strip() if title_col else ''
                price = row.get(price_col) if price_col else None

                if sku and url and 'drive.google.com' in url:
                    results.append({
                        'sku': sku,
                        'gdrive_url': url,
                        'title': title,
                        'price': float(price) if pd.notna(price) else None
                    })
            except (ValueError, TypeError):
                continue

        return results

    def download_image(self, gdrive_url: str, dest_path: str) -> Tuple[bool, str]:
        """
        Descargar imagen de Google Drive.
        Retorna: (success, message)
        """
        download_url = self.gdrive_url_to_download(gdrive_url)
        if not download_url:
            return False, "invalid_url"

        try:
            # Primera solicitud
            session = requests.Session()
            response = session.get(download_url, stream=True, timeout=30)

            # Manejar página de confirmación de descarga grande
            if 'text/html' in response.headers.get('Content-Type', ''):
                # Buscar token de confirmación
                for key, value in response.cookies.items():
                    if key.startswith('download_warning'):
                        download_url = f"{download_url}&confirm={value}"
                        response = session.get(download_url, stream=True, timeout=30)
                        break

            # Verificar que sea imagen
            content_type = response.headers.get('Content-Type', '')
            if 'image' not in content_type and 'octet-stream' not in content_type:
                return False, f"not_image_{content_type[:30]}"

            # Guardar archivo
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            with open(dest_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            # Validar tamaño
            size_kb = os.path.getsize(dest_path) // 1024
            if size_kb < MIN_IMAGE_SIZE_KB:
                os.remove(dest_path)
                return False, f"too_small_{size_kb}kb"

            return True, f"ok_{size_kb}kb"

        except requests.exceptions.Timeout:
            return False, "timeout"
        except requests.exceptions.RequestException as e:
            return False, f"request_error_{str(e)[:50]}"
        except Exception as e:
            return False, f"error_{str(e)[:50]}"

    def batch_download(self, links_list: List[Dict]) -> Dict:
        """
        Descargar batch de imágenes desde GDrive.
        Retorna reporte con estadísticas.
        """
        report = {
            'empresa': self.empresa,
            'timestamp': datetime.now().isoformat(),
            'total': len(links_list),
            'downloaded': 0,
            'failed': 0,
            'skipped_exists': 0,
            'details': []
        }

        print(f"\n[BATCH] Descargando {len(links_list)} imágenes para {self.empresa}...")

        for i, item in enumerate(links_list):
            sku = item['sku']
            url = item['gdrive_url']

            # Determinar extensión (default jpg)
            ext = '.jpg'
            dest_path = os.path.join(self.cache_dir, f"{sku}{ext}")

            # Skip si ya existe
            if os.path.exists(dest_path):
                size_kb = os.path.getsize(dest_path) // 1024
                if size_kb >= MIN_IMAGE_SIZE_KB:
                    report['skipped_exists'] += 1
                    report['details'].append({
                        'sku': sku,
                        'status': 'skipped_exists',
                        'path': dest_path
                    })
                    continue

            # Descargar
            success, message = self.download_image(url, dest_path)

            if success:
                report['downloaded'] += 1
                report['details'].append({
                    'sku': sku,
                    'status': 'downloaded',
                    'message': message,
                    'path': dest_path
                })
                print(f"  [{i+1}/{len(links_list)}] {sku}: OK")
            else:
                report['failed'] += 1
                report['details'].append({
                    'sku': sku,
                    'status': 'failed',
                    'message': message
                })
                print(f"  [{i+1}/{len(links_list)}] {sku}: FAIL - {message}")

            time.sleep(RATE_LIMIT_SECONDS)

            # Checkpoint cada 100
            if (i + 1) % 100 == 0:
                print(f"  [CHECKPOINT {i+1}] downloaded={report['downloaded']}, failed={report['failed']}")

        # Guardar reporte
        report_path = os.path.join(REPORTS_DIR, f"{self.empresa.lower()}_gdrive_download_report.json")
        with open(report_path, 'w') as f:
            json.dump(report, f, indent=2)

        print(f"\n[BATCH COMPLETE] Downloaded: {report['downloaded']}, Failed: {report['failed']}, Skipped: {report['skipped_exists']}")
        return report

    def get_shopify_products_without_image(self) -> List[Dict]:
        """Obtener productos de Shopify que no tienen imagen."""
        if not self.config:
            print(f"[ERROR] No hay config Shopify para {self.empresa}")
            return []

        shop = self.config["shop"]
        if not shop.endswith(".myshopify.com"):
            shop = f"{shop}.myshopify.com"

        token = self.config["token"]
        headers = {"X-Shopify-Access-Token": token}
        products_without_image = []

        url = f"https://{shop}/admin/api/2024-10/products.json"
        params = {"status": "active", "fields": "id,title,handle,variants,images", "limit": 250}

        while url:
            try:
                resp = requests.get(url, headers=headers, params=params, timeout=60)
                if resp.status_code != 200:
                    print(f"[ERROR] Shopify API: {resp.status_code}")
                    break

                data = resp.json()
                for p in data.get("products", []):
                    if len(p.get("images", [])) == 0:
                        sku = p["variants"][0].get("sku", "") if p.get("variants") else ""
                        products_without_image.append({
                            "product_id": p["id"],
                            "sku": sku,
                            "title": p.get("title", ""),
                            "handle": p.get("handle", "")
                        })

                # Paginación
                link_header = resp.headers.get("Link", "")
                url = None
                params = {}
                if 'rel="next"' in link_header:
                    for part in link_header.split(","):
                        if 'rel="next"' in part:
                            url = part.split(";")[0].strip().strip("<>")
                            break

                time.sleep(0.3)
            except Exception as e:
                print(f"[ERROR] {e}")
                break

        print(f"[SHOPIFY] {self.empresa}: {len(products_without_image)} productos sin imagen")
        return products_without_image

    def sync_to_shopify(self, matches: List[Dict] = None) -> Dict:
        """
        Sincronizar imágenes descargadas a Shopify.
        Si no se pasan matches, auto-detecta por SKU exacto.
        """
        if not self.config:
            return {"error": "no_config"}

        shop = self.config["shop"]
        if not shop.endswith(".myshopify.com"):
            shop = f"{shop}.myshopify.com"

        token = self.config["token"]

        # Si no hay matches, construirlos
        if matches is None:
            products = self.get_shopify_products_without_image()
            matches = []

            # Buscar imágenes en cache
            for p in products:
                sku = p['sku'].strip()
                if not sku:
                    continue

                # Buscar imagen por SKU
                for ext in ['.jpg', '.jpeg', '.png', '.webp']:
                    img_path = os.path.join(self.cache_dir, f"{sku}{ext}")
                    if os.path.exists(img_path):
                        matches.append({
                            'product_id': p['product_id'],
                            'sku': sku,
                            'image_path': img_path
                        })
                        break

        report = {
            'empresa': self.empresa,
            'timestamp': datetime.now().isoformat(),
            'total_matches': len(matches),
            'uploaded': 0,
            'failed': 0,
            'details': []
        }

        print(f"\n[SYNC] Subiendo {len(matches)} imágenes a Shopify {self.empresa}...")

        for i, match in enumerate(matches):
            product_id = match['product_id']
            sku = match['sku']
            image_path = match['image_path']

            if not os.path.exists(image_path):
                report['failed'] += 1
                report['details'].append({'sku': sku, 'status': 'file_not_found'})
                continue

            try:
                # Leer y codificar imagen
                with open(image_path, 'rb') as f:
                    image_data = base64.b64encode(f.read()).decode('utf-8')

                url = f"https://{shop}/admin/api/2024-10/products/{product_id}/images.json"
                headers = {
                    "X-Shopify-Access-Token": token,
                    "Content-Type": "application/json"
                }
                payload = {"image": {"attachment": image_data}}

                resp = requests.post(url, headers=headers, json=payload, timeout=60)

                if resp.status_code in [200, 201]:
                    report['uploaded'] += 1
                    report['details'].append({'sku': sku, 'status': 'uploaded'})
                    print(f"  [{i+1}/{len(matches)}] {sku}: UPLOADED")
                else:
                    report['failed'] += 1
                    report['details'].append({'sku': sku, 'status': 'api_error', 'code': resp.status_code})
                    print(f"  [{i+1}/{len(matches)}] {sku}: FAILED - {resp.status_code}")

            except Exception as e:
                report['failed'] += 1
                report['details'].append({'sku': sku, 'status': 'exception', 'error': str(e)[:100]})

            time.sleep(RATE_LIMIT_SECONDS)

        # Guardar reporte
        report_path = os.path.join(REPORTS_DIR, f"{self.empresa.lower()}_gdrive_sync_report.json")
        with open(report_path, 'w') as f:
            json.dump(report, f, indent=2)

        print(f"\n[SYNC COMPLETE] Uploaded: {report['uploaded']}, Failed: {report['failed']}")
        return report

    def full_sync(self, excel_path: str) -> Dict:
        """
        Flujo completo: Excel → GDrive Download → Shopify Upload.
        """
        print(f"\n{'='*60}")
        print(f"GDRIVE FULL SYNC: {self.empresa}")
        print(f"{'='*60}")

        # 1. Extraer links del Excel
        links = self.extract_links_from_excel(excel_path)
        if not links:
            return {"error": "no_links_extracted"}

        # 2. Descargar imágenes
        download_report = self.batch_download(links)

        # 3. Sincronizar a Shopify
        sync_report = self.sync_to_shopify()

        return {
            'empresa': self.empresa,
            'excel_path': excel_path,
            'links_extracted': len(links),
            'download_report': download_report,
            'sync_report': sync_report
        }


def main():
    """CLI para testing."""
    import sys

    if len(sys.argv) < 3:
        print("Uso: python odi_gdrive_ingestor.py <EMPRESA> <EXCEL_PATH>")
        print("Ejemplo: python odi_gdrive_ingestor.py BARA /opt/odi/data/source/BARA/catalogo.xlsx")
        sys.exit(1)

    empresa = sys.argv[1]
    excel_path = sys.argv[2]

    ingestor = GDriveIngestor(empresa)
    result = ingestor.full_sync(excel_path)

    print("\n" + "="*60)
    print("RESULTADO FINAL")
    print("="*60)
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    main()
