#!/usr/bin/env python3
"""
ODI Source Engine V24 — Motor Constitucional
=============================================
ODI lee, procesa, ejecuta, verifica y reporta automaticamente.
Incluye State Registry + Coverage Guard.
Esto no se vuelve a pedir.

Autor: ODI System
Fecha: 25 Feb 2026
"""

import os
import sys
import json
import time
import hashlib
import logging
import re
import requests
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict, field
from concurrent.futures import ThreadPoolExecutor
import traceback

# External libs
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import fitz
except ImportError:
    fitz = None

logging.basicConfig(level=logging.INFO, format='%(asctime)s [SOURCE] %(levelname)s: %(message)s')
log = logging.getLogger("odi-source-engine")

# ============================================================
# CONFIGURACION
# ============================================================
BASE_DATA_DIR = Path("/opt/odi/data")
SOURCE_DIR = BASE_DATA_DIR / "source"
BRANDS_DIR = BASE_DATA_DIR / "brands"
IMAGE_CACHE_DIR = BASE_DATA_DIR / "image_cache"
PROFILES_DIR = BASE_DATA_DIR / "source_profiles"
ACTIONS_DIR = BASE_DATA_DIR / "action_plans"
REPORTS_DIR = BASE_DATA_DIR / "reports"
STATE_DIR = BASE_DATA_DIR / "state"

for d in [SOURCE_DIR, PROFILES_DIR, ACTIONS_DIR, REPORTS_DIR, STATE_DIR]:
    d.mkdir(parents=True, exist_ok=True)

ALL_STORES = ["DFG", "ARMOTOS", "VITTON", "IMBRA", "BARA", "KAIQI", "MCLMOTOS",
              "YOKOMAR", "DUNA", "OH_IMPORTACIONES", "JAPAN", "LEO", "STORE", "VAISAND", "CBI"]
GOVERNED_STORES = {"DFG", "ARMOTOS", "VITTON", "IMBRA", "BARA", "KAIQI", "MCLMOTOS"}
SHOPIFY_RATE_LIMIT = 0.5
GDRIVE_RATE_LIMIT = 0.3

# ============================================================
# DATA CLASSES
# ============================================================
@dataclass
class FileInfo:
    name: str
    hash: str
    mtime: str
    size: int

@dataclass
class SourceSnapshot:
    files: List[Dict] = field(default_factory=list)
    image_links_detected: int = 0
    prices_detected: int = 0

@dataclass
class ShopifySnapshot:
    total_active: int = 0
    total_draft: int = 0
    with_image: int = 0
    without_image: int = 0
    with_ficha: int = 0
    without_ficha: int = 0
    locale: str = "unknown"

@dataclass
class Coverage:
    images_pct: int = 0
    prices_pct: int = 0
    ficha_pct: int = 0
    locale_ok: bool = False

@dataclass
class EnterpriseState:
    empresa: str
    last_scan_hash: str = ""
    last_scan_date: str = ""
    last_execution_date: str = ""
    source_snapshot: Dict = field(default_factory=dict)
    shopify_snapshot: Dict = field(default_factory=dict)
    coverage: Dict = field(default_factory=dict)
    actions_history: List[Dict] = field(default_factory=list)
    snapshot_before: Dict = field(default_factory=dict)
    snapshot_after: Dict = field(default_factory=dict)

# ============================================================
# UTILITIES
# ============================================================
def file_hash(filepath: Path) -> str:
    hash_md5 = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except:
        return ""

def folder_hash(folder: Path) -> str:
    if not folder.exists():
        return ""
    hashes = []
    for f in sorted(folder.rglob("*")):
        if f.is_file():
            hashes.append(f"{f.name}:{f.stat().st_mtime}")
    return hashlib.sha256("|".join(hashes).encode()).hexdigest()[:16]

def load_brand_config(store: str) -> Dict:
    path = BRANDS_DIR / f"{store.lower()}.json"
    try:
        with open(path) as f:
            return json.load(f)
    except:
        return {}

def gdrive_to_download_url(view_url: str) -> Optional[str]:
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', view_url)
    if match:
        return f"https://drive.google.com/uc?export=download&id={match.group(1)}"
    return None

def is_gdrive_url(url: str) -> bool:
    return "drive.google.com" in url or "docs.google.com" in url

# ============================================================
# SOURCE ENGINE CLASS
# ============================================================
class SourceEngine:
    def __init__(self):
        self.states: Dict[str, EnterpriseState] = {}
    
    # ------ STATE MANAGEMENT ------
    def _load_state(self, empresa: str) -> EnterpriseState:
        path = STATE_DIR / f"{empresa}_state.json"
        if path.exists():
            with open(path) as f:
                data = json.load(f)
                return EnterpriseState(**data)
        return EnterpriseState(empresa=empresa)
    
    def _save_state(self, empresa: str, state: EnterpriseState):
        path = STATE_DIR / f"{empresa}_state.json"
        with open(path, "w") as f:
            json.dump(asdict(state), f, indent=2, default=str)
    
    def _update_coverage(self, state: EnterpriseState) -> Dict:
        src = state.source_snapshot
        shop = state.shopify_snapshot
        total = shop.get("total_active", 0) or 1
        coverage = {
            "images_pct": int((shop.get("with_image", 0) / total) * 100) if total else 0,
            "prices_pct": 100 if shop.get("total_active", 0) > 0 else 0,
            "ficha_pct": int((shop.get("with_ficha", 0) / total) * 100) if total else 0,
            "locale_ok": shop.get("locale", "") == "es"
        }
        state.coverage = coverage
        return coverage
    
    def _should_execute(self, state: EnterpriseState) -> Tuple[bool, List[str]]:
        reasons = []
        cov = state.coverage
        src = state.source_snapshot
        shop = state.shopify_snapshot
        
        if cov.get("images_pct", 0) < 100 and src.get("image_links_detected", 0) > 0:
            reasons.append("images_pct < 100 con fuentes disponibles")
        if cov.get("ficha_pct", 0) < 100:
            reasons.append("ficha_pct < 100")
        if not cov.get("locale_ok", False):
            reasons.append("locale no es español")
        if shop.get("with_image", 0) < src.get("image_links_detected", 0):
            reasons.append("shopify images < source links")
        
        return len(reasons) > 0, reasons
    
    # ------ PHASE 1: DEEP SCAN ------
    def deep_scan_all(self) -> Dict:
        log.info("=" * 60)
        log.info("DEEP SCAN ALL")
        log.info("=" * 60)
        results = {}
        for store in ALL_STORES:
            try:
                results[store] = self.deep_scan(store)
            except Exception as e:
                log.error(f"Error {store}: {e}")
        self._print_summary(results)
        return results
    
    def deep_scan(self, empresa: str) -> Dict:
        log.info(f"[{empresa}] Deep scan...")
        state = self._load_state(empresa)
        
        search_paths = [
            SOURCE_DIR / empresa,
            BASE_DATA_DIR / empresa,
            IMAGE_CACHE_DIR / f"{empresa}_GDRIVE",
            IMAGE_CACHE_DIR / f"{empresa}_pages",
        ]
        
        files_info = []
        total_links = 0
        total_prices = 0
        
        for sp in search_paths:
            if not sp.exists():
                continue
            for fp in sp.rglob("*"):
                if fp.is_file():
                    ext = fp.suffix.lower()
                    info = {"name": fp.name, "hash": file_hash(fp)[:8], 
                            "mtime": datetime.fromtimestamp(fp.stat().st_mtime).isoformat(),
                            "size": fp.stat().st_size, "type": ext}
                    files_info.append(info)
                    
                    # Analyze content
                    if ext == ".json" and ("gdrive" in fp.name or "mapping" in fp.name):
                        try:
                            with open(fp) as f:
                                data = json.load(f)
                                if isinstance(data, list):
                                    for item in data:
                                        if item.get("gdrive_url") or item.get("image_url"):
                                            total_links += 1
                                        if item.get("price"):
                                            total_prices += 1
                        except:
                            pass
                    
                    elif ext in [".xlsx", ".xls"] and openpyxl:
                        try:
                            wb = openpyxl.load_workbook(fp, data_only=False)
                            sheet = wb.active
                            for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 2000)):
                                for cell in row:
                                    if cell.hyperlink or (cell.value and "drive.google" in str(cell.value)):
                                        total_links += 1
                                    if isinstance(cell.value, (int, float)) and cell.value > 100:
                                        total_prices += 1
                            wb.close()
                        except:
                            pass
        
        # Also check loose mappings
        for loose in BASE_DATA_DIR.glob(f"{empresa}_*.json"):
            if "mapping" in loose.name or "gdrive" in loose.name:
                try:
                    with open(loose) as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            for item in data:
                                if item.get("gdrive_url") or item.get("image_url"):
                                    total_links += 1
                except:
                    pass
        
        state.source_snapshot = {
            "files": files_info,
            "image_links_detected": total_links,
            "prices_detected": total_prices
        }
        state.last_scan_hash = folder_hash(SOURCE_DIR / empresa)
        state.last_scan_date = datetime.now(timezone.utc).isoformat()
        
        self._save_state(empresa, state)
        
        return {"files": len(files_info), "links": total_links, "prices": total_prices}
    
    # ------ PHASE 2: ACTION PLAN ------
    def generate_action_plan(self, empresa: str) -> Dict:
        log.info(f"[{empresa}] Generating action plan...")
        state = self._load_state(empresa)
        src = state.source_snapshot
        
        actions = []
        
        if src.get("image_links_detected", 0) > 0:
            actions.append({"type": "DOWNLOAD_IMAGES_FROM_LINKS", "priority": "HIGH",
                           "count": src["image_links_detected"]})
        
        if src.get("prices_detected", 0) > 0:
            actions.append({"type": "VERIFY_AND_UPDATE_PRICES", "priority": "HIGH",
                           "count": src["prices_detected"]})
        
        if empresa in GOVERNED_STORES:
            actions.append({"type": "GENERATE_FICHA_7_CUERPOS", "priority": "MEDIUM"})
            actions.append({"type": "FIX_LOCALE", "priority": "LOW"})
        
        plan = {"empresa": empresa, "date": datetime.now(timezone.utc).isoformat(), "actions": actions}
        
        path = ACTIONS_DIR / f"{empresa}_actions.json"
        with open(path, "w") as f:
            json.dump(plan, f, indent=2)
        
        return plan
    
    # ------ PHASE 3: EXECUTE ------
    def execute_plan(self, empresa: str, dry_run: bool = False) -> Dict:
        log.info(f"[{empresa}] Executing plan (dry_run={dry_run})...")
        state = self._load_state(empresa)
        
        # Snapshot BEFORE
        state.snapshot_before = self._get_shopify_snapshot(empresa)
        
        results = {"empresa": empresa, "dry_run": dry_run, "actions": {}}
        
        plan_path = ACTIONS_DIR / f"{empresa}_actions.json"
        if not plan_path.exists():
            self.generate_action_plan(empresa)
        
        with open(plan_path) as f:
            plan = json.load(f)
        
        for action in plan.get("actions", []):
            atype = action.get("type")
            try:
                if atype == "DOWNLOAD_IMAGES_FROM_LINKS":
                    results["actions"][atype] = self._exec_download_images(empresa, dry_run)
                elif atype == "VERIFY_AND_UPDATE_PRICES":
                    results["actions"][atype] = self._exec_update_prices(empresa, dry_run)
                elif atype == "FIX_LOCALE":
                    results["actions"][atype] = {"status": "pending"}
                elif atype == "GENERATE_FICHA_7_CUERPOS":
                    results["actions"][atype] = {"status": "pending"}
            except Exception as e:
                results["actions"][atype] = {"error": str(e)}
        
        state.last_execution_date = datetime.now(timezone.utc).isoformat()
        state.actions_history.append({
            "date": state.last_execution_date,
            "actions": list(results["actions"].keys()),
            "dry_run": dry_run
        })
        
        self._save_state(empresa, state)
        return results
    
    def _exec_download_images(self, empresa: str, dry_run: bool) -> Dict:
        result = {"downloaded": 0, "skipped": 0, "failed": 0, "uploaded": 0}
        
        # Find links file
        links = []
        for src in [SOURCE_DIR / empresa / "gdrive_links.json",
                    SOURCE_DIR / empresa / "pending_gdrive_download.json",
                    BASE_DATA_DIR / f"{empresa}_image_mapping.json"]:
            if src.exists():
                with open(src) as f:
                    links = json.load(f)
                break
        
        if not links:
            return result
        
        cache_dir = IMAGE_CACHE_DIR / f"{empresa}_GDRIVE"
        cache_dir.mkdir(parents=True, exist_ok=True)
        
        brand_config = load_brand_config(empresa)
        shop_cfg = brand_config.get("shopify", {})
        
        for item in links[:500]:  # Limit per run
            sku = item.get("sku", "")
            url = item.get("gdrive_url") or item.get("image_url", "")
            product_id = item.get("product_id")
            
            if not url or not sku:
                continue
            
            local_path = cache_dir / f"{sku}.jpg"
            if local_path.exists():
                result["skipped"] += 1
                continue
            
            if dry_run:
                result["downloaded"] += 1
                continue
            
            try:
                dl_url = gdrive_to_download_url(url) if is_gdrive_url(url) else url
                if not dl_url:
                    result["failed"] += 1
                    continue
                
                resp = requests.get(dl_url, timeout=30, allow_redirects=True)
                if resp.status_code == 200 and len(resp.content) > 5000:
                    with open(local_path, "wb") as f:
                        f.write(resp.content)
                    result["downloaded"] += 1
                    
                    if product_id and shop_cfg.get("shop"):
                        if self._upload_to_shopify(shop_cfg, product_id, local_path):
                            result["uploaded"] += 1
                    
                    time.sleep(GDRIVE_RATE_LIMIT)
                else:
                    result["failed"] += 1
            except Exception as e:
                log.warning(f"Download error {sku}: {e}")
                result["failed"] += 1
        
        return result
    
    def _upload_to_shopify(self, cfg: Dict, product_id: int, img_path: Path) -> bool:
        import base64
        shop, token = cfg.get("shop"), cfg.get("token")
        if not shop or not token:
            return False
        try:
            with open(img_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            url = f"https://{shop}/admin/api/2024-10/products/{product_id}/images.json"
            headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
            resp = requests.post(url, headers=headers, json={"image": {"attachment": b64}}, timeout=30)
            time.sleep(SHOPIFY_RATE_LIMIT)
            return resp.status_code in [200, 201]
        except:
            return False
    
    def _exec_update_prices(self, empresa: str, dry_run: bool) -> Dict:
        return {"checked": 0, "updated": 0, "status": "pending"}
    
    # ------ PHASE 4: VERIFY ------
    def verify(self, empresa: str) -> Dict:
        log.info(f"[{empresa}] Verifying Shopify state...")
        state = self._load_state(empresa)
        
        snapshot = self._get_shopify_snapshot(empresa)
        state.shopify_snapshot = snapshot
        state.snapshot_after = snapshot
        
        self._update_coverage(state)
        self._save_state(empresa, state)
        
        return snapshot
    
    def _get_shopify_snapshot(self, empresa: str) -> Dict:
        brand_cfg = load_brand_config(empresa)
        shop_cfg = brand_cfg.get("shopify", {})
        shop, token = shop_cfg.get("shop"), shop_cfg.get("token")
        
        snap = {"total_active": 0, "total_draft": 0, "with_image": 0, 
                "without_image": 0, "with_ficha": 0, "locale": "unknown"}
        
        if not shop or not token:
            return snap
        
        headers = {"X-Shopify-Access-Token": token}
        try:
            # Count active
            r = requests.get(f"https://{shop}/admin/api/2024-10/products/count.json?published_status=published",
                           headers=headers, timeout=10)
            if r.status_code == 200:
                snap["total_active"] = r.json().get("count", 0)
            
            # Count draft
            r = requests.get(f"https://{shop}/admin/api/2024-10/products/count.json?published_status=unpublished",
                           headers=headers, timeout=10)
            if r.status_code == 200:
                snap["total_draft"] = r.json().get("count", 0)
            
            # Sample for images
            r = requests.get(f"https://{shop}/admin/api/2024-10/products.json?limit=50&fields=id,images,body_html",
                           headers=headers, timeout=15)
            if r.status_code == 200:
                products = r.json().get("products", [])
                for p in products:
                    if p.get("images"):
                        snap["with_image"] += 1
                    else:
                        snap["without_image"] += 1
                    body = p.get("body_html", "") or ""
                    if "<table" in body.lower() or "ficha" in body.lower():
                        snap["with_ficha"] += 1
        except Exception as e:
            log.warning(f"Shopify error {empresa}: {e}")
        
        return snap
    
    # ------ PHASE 5: REPORT ------
    def report(self, empresa: str) -> Dict:
        log.info(f"[{empresa}] Generating report...")
        state = self._load_state(empresa)
        
        src = state.source_snapshot
        shop = state.shopify_snapshot
        cov = state.coverage
        
        diff = {}
        if state.snapshot_before and state.snapshot_after:
            for k in ["total_active", "with_image"]:
                before = state.snapshot_before.get(k, 0)
                after = state.snapshot_after.get(k, 0)
                if before != after:
                    diff[k] = {"before": before, "after": after, "delta": after - before}
        
        report = {
            "empresa": empresa,
            "report_date": datetime.now(timezone.utc).isoformat(),
            "sources_detected": {
                "total_files": len(src.get("files", [])),
                "image_links": src.get("image_links_detected", 0),
                "prices": src.get("prices_detected", 0)
            },
            "shopify_state": shop,
            "coverage": cov,
            "diff": diff,
            "actions_history": state.actions_history[-5:]
        }
        
        path = REPORTS_DIR / f"{empresa}_source_engine_report.json"
        with open(path, "w") as f:
            json.dump(report, f, indent=2)
        
        return report
    
    # ------ PHASE 6: AUTO RUN ------
    def auto_run(self, force: bool = False):
        log.info("=" * 60)
        log.info("AUTO RUN — Coverage Guard Active")
        log.info("=" * 60)
        
        # 1. Deep scan all
        self.deep_scan_all()
        
        # 2. Verify all (get current Shopify state)
        for empresa in ALL_STORES:
            try:
                self.verify(empresa)
            except:
                pass
        
        # 3. Generate plans and check coverage
        for empresa in GOVERNED_STORES:
            try:
                state = self._load_state(empresa)
                self._update_coverage(state)
                self._save_state(empresa, state)
                
                should_exec, reasons = self._should_execute(state)
                
                if should_exec or force:
                    log.info(f"[{empresa}] Ejecutando por: {reasons}")
                    self.generate_action_plan(empresa)
                    self.execute_plan(empresa, dry_run=False)
                    self.verify(empresa)
                else:
                    log.info(f"[{empresa}] Coverage OK, skip execution")
                
                self.report(empresa)
            except Exception as e:
                log.error(f"Error {empresa}: {e}")
        
        # 4. Report legacy stores
        for empresa in set(ALL_STORES) - GOVERNED_STORES:
            try:
                self.report(empresa)
            except:
                pass
        
        log.info("AUTO RUN completado")
    
    def _print_summary(self, results: Dict):
        log.info("\n" + "=" * 60)
        log.info("SCAN SUMMARY")
        log.info("=" * 60)
        for emp, data in results.items():
            status = []
            if data.get("links", 0) > 0:
                status.append(f"LINKS:{data['links']}")
            if data.get("prices", 0) > 0:
                status.append(f"PRICES:{data['prices']}")
            if data.get("files", 0) > 0:
                status.append(f"FILES:{data['files']}")
            log.info(f"  {emp:20} | {' | '.join(status) or 'EMPTY'}")

# ============================================================
# CLI
# ============================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description="ODI Source Engine V24")
    parser.add_argument("command", choices=["scan", "plan", "execute", "verify", "report", "auto", "status"])
    parser.add_argument("--empresa", "-e", help="Empresa especifica")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true", help="Forzar ejecucion aunque coverage OK")
    
    args = parser.parse_args()
    engine = SourceEngine()
    
    if args.command == "scan":
        if args.empresa:
            print(json.dumps(engine.deep_scan(args.empresa.upper()), indent=2))
        else:
            engine.deep_scan_all()
    
    elif args.command == "plan":
        empresas = [args.empresa.upper()] if args.empresa else ALL_STORES
        for e in empresas:
            print(json.dumps(engine.generate_action_plan(e), indent=2))
    
    elif args.command == "execute":
        empresas = [args.empresa.upper()] if args.empresa else list(GOVERNED_STORES)
        for e in empresas:
            print(json.dumps(engine.execute_plan(e, dry_run=args.dry_run), indent=2))
    
    elif args.command == "verify":
        empresas = [args.empresa.upper()] if args.empresa else ALL_STORES
        for e in empresas:
            r = engine.verify(e)
            print(f"{e}: active={r['total_active']}, draft={r['total_draft']}, img={r['with_image']}")
    
    elif args.command == "report":
        empresas = [args.empresa.upper()] if args.empresa else ALL_STORES
        for e in empresas:
            engine.report(e)
    
    elif args.command == "auto":
        engine.auto_run(force=args.force)
    
    elif args.command == "status":
        for e in ALL_STORES:
            state = engine._load_state(e)
            cov = state.coverage
            src = state.source_snapshot
            log.info(f"{e:20} | img:{cov.get('images_pct',0):3}% | ficha:{cov.get('ficha_pct',0):3}% | links:{src.get('image_links_detected',0):4}")

if __name__ == "__main__":
    main()
